#!/usr/bin/env python3
"""
JDBC Warehouse Test Skill - XLSX 生成器
生成符合 JDBC 批量入仓接口规范的 23 列 xlsx 文件
"""
import os
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import config
from table_reader import TableReader


class XlsxGenerator:
    """XLSX 文件生成器"""

    def __init__(self):
        self.reader = TableReader()

    def generate(self, database, table, scenario='success', env=None, output_path=None, **kwargs):
        """
        生成 XLSX 测试文件

        Args:
            database: 数据库名
            table: 表名
            scenario: 测试场景（success, failed_F001等）
            env: 环境名（可选）
            output_path: 输出路径（可选）
            **kwargs: 其他参数

        Returns:
            dict: {'success': bool, 'file_path': str, 'message': str}
        """
        try:
            print(f"\n{'='*60}")
            print("JDBC 入仓测试文件生成器")
            print(f"{'='*60}")
            print(f"数据库: {database}")
            print(f"表名: {table}")
            print(f"场景: {config.SCENARIOS.get(scenario, scenario)}")
            print(f"{'='*60}\n")

            # 1. 读取表结构信息
            print("正在读取表结构...")
            table_info = self.reader.get_table_info(database, table, env=env)
            print(f"✓ 表结构读取成功")
            print(f"  实例: {table_info['instance']}")
            print(f"  数据源类型: {table_info['db_type']}")
            print(f"  主键: {table_info['primary_key']}")
            print(f"  字段数: {len(table_info['columns'])}")

            # 2. 检查时间字段
            time_fields = self.reader.has_time_fields(table_info)
            print(f"  时间字段: {time_fields['created_field']}, {time_fields['updated_field']}")

            # 3. 验证元数据（如果是 failed_F004 场景，跳过此检查）
            if scenario != 'failed_F004':
                print("\n正在验证元数据...")
                metadata_exists = self.reader.verify_metadata_exists(
                    table_info['instance'],
                    table_info['database'],
                    table_info['table']
                )
                if not metadata_exists:
                    print("⚠️  警告: 元数据未完善，建议先执行 metadata-complete")
                else:
                    print("✓ 元数据已完善")

            # 4. 生成 23 列数据
            print("\n正在生成测试数据...")
            row_data = self._generate_row_data(table_info, scenario, time_fields)

            # 5. 创建 Excel 文件
            print("正在创建 Excel 文件...")
            wb = Workbook()
            ws = wb.active
            ws.title = "批量任务"

            # 写入表头
            for col_idx, header in enumerate(config.EXCEL_HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入数据行
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=2, column=col_idx, value=value)

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # 6. 保存文件
            if output_path is None:
                output_path = config.get_output_path()

            os.makedirs(output_path, exist_ok=True)

            filename = config.generate_filename(scenario, table)
            file_path = os.path.join(output_path, filename)

            wb.save(file_path)
            print(f"✓ 文件保存成功")

            print(f"\n{'='*60}")
            print("✅ 测试文件生成完成")
            print(f"{'='*60}")
            print(f"文件路径: {file_path}")
            print(f"文件名称: {filename}")
            print(f"{'='*60}\n")

            return {
                'success': True,
                'file_path': file_path,
                'filename': filename,
                'message': '测试文件生成成功'
            }

        except Exception as e:
            error_msg = f"生成失败: {str(e)}"
            print(f"\n❌ {error_msg}\n")
            return {
                'success': False,
                'file_path': None,
                'message': error_msg
            }

    def _generate_row_data(self, table_info, scenario, time_fields):
        """
        生成一行数据（23 列）

        Args:
            table_info: 表信息
            scenario: 测试场景
            time_fields: 时间字段信息

        Returns:
            list: 23 列数据
        """
        database = table_info['database']
        table = table_info['table']
        instance = table_info['instance']
        primary_key = table_info['primary_key']

        # 推断数据源类型
        db_type = config.infer_db_type(database)

        # 获取抽数/处理方式
        extract_process = config.EXTRACT_PROCESS_RULES.get(scenario, config.EXTRACT_PROCESS_RULES['success'])

        # 生成随机调度时间
        hour = random.randint(
            config.SCHEDULE_TIME_RANGE['hour_start'],
            config.SCHEDULE_TIME_RANGE['hour_end']
        )
        minute = random.choice(range(0, 60, config.SCHEDULE_TIME_RANGE['minute_step']))
        schedule_time = f"{hour:02d}:{minute:02d}"

        # 随机选择用户旅程节点
        journey_node = random.choice(config.USER_JOURNEY_NODES)

        # 生成抽数数据源名称
        datasource_name = config.generate_datasource_name(db_type, instance, database)

        # 构造 23 列数据
        row_data = [
            db_type,                                    # 1. 数据源类型
            instance.lower(),                           # 2. 实例
            database.lower(),                           # 3. 库
            table.lower(),                              # 4. 表
            config.FIXED_VALUES['person_name'],         # 5. 业务负责人
            config.FIXED_VALUES['person_uid'],          # 6. 业务负责人UID
            config.FIXED_VALUES['person_name'],         # 7. 技术负责人
            config.FIXED_VALUES['person_uid'],          # 8. 技术负责人UID
            config.FIXED_VALUES['person_name'],         # 9. 创建人
            config.FIXED_VALUES['person_uid'],          # 10. 创建人UID
            journey_node,                               # 11. 用户旅程节点
            config.FIXED_VALUES['purpose'],             # 12. 需求目的
            extract_process['extract'],                 # 13. 抽数方式
            extract_process['process'],                 # 14. 处理方式
            config.FIXED_VALUES['schedule_cycle'],      # 15. 调度周期
            schedule_time,                              # 16. 调度时间
            config.FIXED_VALUES['cpu'],                 # 17. CPU
            config.FIXED_VALUES['memory'],              # 18. 内存
            time_fields['created_field'],               # 19. 创建时间字段
            time_fields['updated_field'],               # 20. 更新时间字段
            config.FIXED_VALUES['batch_size'],          # 21. 批量条数
            datasource_name,                            # 22. 抽数数据源
            primary_key                                 # 23. 抽数主键
        ]

        # 根据异常场景调整数据
        row_data = self._apply_failed_scenario(row_data, scenario)

        return row_data

    def _apply_failed_scenario(self, row_data, scenario):
        """
        根据异常场景调整数据

        Args:
            row_data: 原始行数据
            scenario: 场景类型

        Returns:
            list: 调整后的行数据
        """
        if scenario == 'failed_F001':
            # 字段缺失：删除最后一列（抽数主键）
            return row_data[:-1]

        elif scenario == 'failed_F002':
            # 字段格式错误：调度时间格式错误
            row_data[15] = "25:99"  # 列16: 调度时间

        elif scenario == 'failed_F003':
            # 联动规则冲突：已在 EXTRACT_PROCESS_RULES 中定义

            pass

        elif scenario == 'failed_F004':
            # 元数据未完善：数据正常，但表未完善元数据
            pass

        return row_data


def main():
    """测试函数"""
    import sys

    if len(sys.argv) < 3:
        print("用法: python xlsx_generator.py <database> <table> [scenario] [env]")
        print("\n场景类型:")
        for key, value in config.SCENARIOS.items():
            print(f"  {key}: {value}")
        sys.exit(1)

    database = sys.argv[1]
    table = sys.argv[2]
    scenario = sys.argv[3] if len(sys.argv) > 3 else 'success'
    env = sys.argv[4] if len(sys.argv) > 4 else None

    generator = XlsxGenerator()
    result = generator.generate(database, table, scenario, env=env)

    if result['success']:
        print(f"✅ 成功: {result['message']}")
        print(f"📄 文件: {result['file_path']}")
        sys.exit(0)
    else:
        print(f"❌ 失败: {result['message']}")
        sys.exit(1)


if __name__ == '__main__':
    main()
